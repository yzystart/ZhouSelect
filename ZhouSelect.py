import random
import sys
from time import sleep
from openpyxl import Workbook
from selenium import webdriver
from lxml import etree
import os
import requests
import json
import openpyxl

# 全局配置
config = {}

# excel生成模板
template = ['客户名称', '经营状态', '工商注册地址', '法人姓名', '工商联系方式', '查询到的机构']

# 请求头
query_headers = {}

# 初始化请求头
def init_cookie(Cookie):
    global query_headers
    query_headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Host': 'www.tianyancha.com',
        'Referer': 'https://www.tianyancha.com/',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
        'sec-ch-ua-mobile': '?0',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'Cookie': Cookie,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.72 Safari/537.36'
    }

# 初始化excel模板
def init_excel_template():
    print('>正在生成模板...')
    try:
        work_book=Workbook()
        work_sheet = work_book.active
        work_sheet.title='Sheet1'
        work_sheet.append(template)
    except Exception as e:
        work_book.save('template.xlsx')
        work_book.close()
    finally:
        work_book.save('template.xlsx')
        work_book.close()
        input('>模板生成完成,回车结束程序')



# 生成配置文件
def init_config():
    global config
    config = {
        # 当前操作的excel
        'excel': 'template.xlsx',
        # 工作表单
        'work_sheet': 'Sheet1',
        # 上一次录入到的行
        'last_line': '1'
    }
    savefile()

# 保存配置文件
def savefile():
    with open(os.path.abspath('.') + '/config.json', encoding='utf-8', mode='w') as writefile:
        writefile.write(json.dumps(config, ensure_ascii=False))

# 读取配置文件
def read_config():
    global config
    try:
        print('正在读取配置文件...')
        # 读取配置文件
        with open(os.path.abspath('.') + '/config.json', encoding='utf-8-sig', mode='r') as readfile:
            # 读取文本
            str_config = readfile.read()
            # 转化为字典
            config = json.loads(str_config)
            print(config)
    except FileNotFoundError:
        print('没有配置文件,开始生成..')
        init_config()
        input('生成结束,请重新打开程序,回车后退出程序！')
        sys.exit()

# 主查询方法
def query(org):
    result={
        'query_org_name': org,
        'user': '无',
        'status': '未查询到',
        'phone': '无',
        'address': '无',
        'query_org_result_name': '未查询到'
    }
    query_url = "https://www.tianyancha.com/search?key="+org
    query_response = requests.get(url=query_url,headers=query_headers)
    html = etree.HTML(query_response.text)
    try:
        # check_login()
        # 所有搜索列表
        resule_list_div=html.xpath("//div[@class='result-list sv-search-container']")[0]
        # 一般第一个就是正确的，取搜索列表中的第一个
        content_div=resule_list_div.xpath("./div/div/div[@class='content']/div/a/em[text()='"+org+"']/../../..")[0]
        print('***************************************')
        try:
            org_name=content_div.xpath('./../div[@class="triangle-xcx"]/div[@class="xcx-block"]/div[@class="info"]')[0].text
            # org_name=resule_list_div.xpath('./div/div[@class="search-result-single  "]/div[@class="triangle-xcx"]/div[@class="xcx-block"]/div[@class="info"]')[0].text
            print('公司名->',org_name)
            result['query_org_result_name']=org_name
        except Exception as e:
            print('没有定位到公司名',e)
        try:
            user=content_div.xpath('./div[@class="info row text-ellipsis"]/div[1]/a')[0].text
            print('法人/经营者->',user)
            result['user']=user
        except Exception as e:
            print('没有查询到法人/经营者',e)
        try:
            status= content_div.xpath("./div[@class='header']/div")[0].text
            print('经营状态 ->',status)
            result['status']=status
        except Exception as e:
            print('没有查询到经营状态',e)
        try:
            phone = content_div.xpath('.//span[text()="电话："]/../span[2]/span')[0].text
            print('电话 ->',phone)
            result['phone']=phone
        except Exception as e:
            print('没有查询到电话',e)
        try:
            address = content_div.xpath('.//span[text()="地址："]/../span[2]')[0].text
            print('地址 ->',address)
            result['address']=address
        except Exception as e:
            print('没有查询到地址',e)
        print('***************************************')
    except Exception as e:
        print('没有查询到该公司：'+org,e)
    return result

# excel 录入方法
def do_excel():
    read_config()
    print('>开始加载excel...')
    work_book = openpyxl.load_workbook(config['excel'])
    print('>开始加载表单...')
    work_sheet = work_book[config['work_sheet']]
    print('>开始读取总条数...')
    work_len = work_sheet.max_row
    print('>开始读取进度...')
    progress = config['last_line']
    print('>工作表单加载完成->', config['work_sheet'])
    print('>工作excel加载完成->', config['excel'])
    print('> size ->', str(work_len))
    print('>上一次执行到->', progress)
    print('现在开始执行自动录入,请不要关闭程序,请手动备份文件...')
    try:
        for index, row in enumerate(work_sheet.rows):
            if index < int(config['last_line']):
                print('已略过！index：', index)
                continue
            print('当前下标', index)
            instrt = query(str(work_sheet.cell(index + 1, column=1).value))
            work_sheet.cell(row=index + 1, column=2, value=instrt['status'])
            work_sheet.cell(row=index + 1, column=3, value=instrt['address'])
            work_sheet.cell(row=index + 1, column=4, value=instrt['user'])
            work_sheet.cell(row=index + 1, column=5, value=instrt['phone'])
            work_sheet.cell(row=index + 1, column=6, value=instrt['query_org_result_name'])
            config['last_line'] = str(index)
            savefile()
            work_book.save(config['excel'])
            if index % 100 == 0:
                work_book.save('自动备份_' + str(index) + '.xlsx')
            sleep_time = random.randint(1, 5)
            print('>反反爬睡眠：' + str(sleep_time) + '秒')
            sleep(sleep_time)
    except Exception as e:
        print('程序发生了意想不到的错误！', e)
    finally:
        savefile()
        work_book.save(config['excel'])
        work_book.close()
    input('end...')



print('***************************************')
print('**********欢迎来到老周查询器2.0***********')
print('***************************************')
print('*************选择一个选项？**************')
print('***************************************')
print('**************1.单查询*****************')
print('**************2.Excel录入**************')
print('**************3.生成Excel模板***********')
print('***************************************')
print('>当前的操作路径为', os.path.abspath('.'))

select=input()

if select=='2':
    read_config()
# 驱动路径
bro = webdriver.Chrome(executable_path=os.path.abspath('.') + '\\' + "chromedriver.exe")
# 打开网页获取cookie
bro.get("https://www.tianyancha.com/")
#
input('请在页面登录你的天眼查账号，登录完成后回到cmd窗口回车')
Cookie=''
# 设置cookie
print('>开始读取Cookie...')
for cookie in bro.get_cookies():
    print(cookie['name']+'='+cookie['value']+';')
    Cookie=Cookie+cookie['name']+'='+cookie['value']+';'
Cookie=Cookie[0:-1]
init_cookie(Cookie)

# 入口
if __name__ == "__main__":

    if select=='1':
        while True:
            try:
                query(input('请输入一个机构：'))
            except Exception as e:
                print(e)
                continue
    if select =='2':do_excel()
    if select == '3':init_excel_template()





